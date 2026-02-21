"""Excel 报告生成器 — 生成串标围标分析报告 (.xlsx)"""
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# === Style Presets ===
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
SECTION_FONT = Font(name="微软雅黑", size=10, bold=True, color="2E75B6")
BODY_FONT = Font(name="微软雅黑", size=10, color="333333")
DETAIL_FONT = Font(name="微软雅黑", size=9, color="555555")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

RISK_COLORS = {
    "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "high": PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
    "medium": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "low": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
}
RISK_NAMES = {"critical": "严重", "high": "高风险", "medium": "中等", "low": "低风险"}
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

TYPE_NAMES = {
    "content_similarity": "文本相似度", "metadata_match": "元数据关联",
    "format_match": "格式指纹", "timestamp_cluster": "时间戳聚集",
    "entity_cross": "实体交叉", "error_pattern": "错误模式", "price_analysis": "报价分析",
}


def _safe(val, default=""):
    return val if val is not None else default


def _write_row(ws, row, values, fonts=None, fills=None, aligns=None, border=True):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = fonts[col-1] if fonts and col <= len(fonts) else BODY_FONT
        if fills and col <= len(fills) and fills[col-1]:
            cell.fill = fills[col-1]
        cell.alignment = aligns[col-1] if aligns and col <= len(aligns) else CENTER
        if border:
            cell.border = THIN_BORDER


class ExcelReportGenerator:

    @staticmethod
    def generate(project: Dict, documents: List[Dict], results: List[Dict],
                 risk_summary: Dict) -> io.BytesIO:
        wb = Workbook()
        ExcelReportGenerator._build_overview_sheet(wb, project, risk_summary)
        ExcelReportGenerator._build_documents_sheet(wb, documents)
        ExcelReportGenerator._build_results_sheet(wb, results)
        ExcelReportGenerator._build_detail_sheet(wb, results)
        ExcelReportGenerator._build_alerts_sheet(wb, results)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ==================== Sheet 1: 项目概览 ====================
    @staticmethod
    def _build_overview_sheet(wb, project, risk_summary):
        ws = wb.create_sheet("项目概览", 0)
        ws.merge_cells("A1:F1")
        ws["A1"] = "串标围标分析报告"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 40
        ws.merge_cells("A2:F2")
        ws["A2"] = f"报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
        ws["A2"].alignment = CENTER

        info_rows = [
            ("项目名称", _safe(project.get("name"))),
            ("项目编号", _safe(project.get("project_code"), "N/A")),
            ("文档数量", str(project.get("document_count", 0))),
            ("综合风险评分", f"{(project.get('risk_score') or 0.0):.1f} / 100"),
            ("风险等级", RISK_NAMES.get(project.get("risk_level") or "low", "低风险")),
        ]
        row = 4
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "项目基本信息"
        ws[f"A{row}"].font = SUBTITLE_FONT
        row += 1
        for label, value in info_rows:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(name="微软雅黑", size=10, bold=True)
            ws.merge_cells(f"B{row}:F{row}")
            ws[f"B{row}"] = value
            ws[f"B{row}"].font = BODY_FONT
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "各维度检测得分"
        ws[f"A{row}"].font = SUBTITLE_FONT
        row += 1
        for i, h in enumerate(["检测维度", "得分", "风险等级"]):
            c = ws.cell(row=row, column=i+1, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        row += 1
        dimension_scores = risk_summary.get("dimension_scores", {})
        for dim_key, dim_name in TYPE_NAMES.items():
            score = dimension_scores.get(dim_key, 0.0)
            risk = "critical" if score >= 0.7 else "high" if score >= 0.5 else "medium" if score >= 0.3 else "low"
            ws.cell(row=row, column=1, value=dim_name).font = BODY_FONT
            ws.cell(row=row, column=2, value=f"{score:.1%}").font = BODY_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            rc = ws.cell(row=row, column=3, value=RISK_NAMES[risk])
            rc.fill = RISK_COLORS[risk]
            rc.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF" if risk in ["critical","high"] else "333333")
            rc.alignment = CENTER
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
        ws.column_dimensions["A"].width = 18
        for col in "BCDEF":
            ws.column_dimensions[col].width = 16

    # ==================== Sheet 2: 文档清单 ====================
    @staticmethod
    def _build_documents_sheet(wb, documents):
        ws = wb.create_sheet("文档清单")
        headers = ["序号", "投标单位", "文件名", "文件类型", "大小(KB)", "作者", "创建软件", "创建时间", "页数", "状态"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        for idx, doc in enumerate(documents, 1):
            vals = [idx, _safe(doc.get("company_name")), _safe(doc.get("file_name")), _safe(doc.get("file_type")),
                    round(doc.get("file_size", 0) / 1024, 1), _safe(doc.get("meta_author")),
                    _safe(doc.get("meta_creator")), _safe(doc.get("meta_created_time")),
                    doc.get("page_count", 0),
                    "已解析" if doc.get("parsed") == 1 else "失败" if doc.get("parsed") == 2 else "未解析"]
            _write_row(ws, idx + 1, vals)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(headers[col-1])) * 2)

    # ==================== Sheet 3: 检测结果明细 ====================
    @staticmethod
    def _build_results_sheet(wb, results):
        ws = wb.create_sheet("检测结果明细")
        headers = ["序号", "检测类型", "单位A", "单位B", "得分", "风险", "说明"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        for idx, r in enumerate(sorted(results, key=lambda x: x.get("score", 0), reverse=True), 1):
            risk = r.get("risk_level") or "low"
            vals = [idx, TYPE_NAMES.get(r.get("analysis_type",""), ""),
                    _safe(r.get("company_a")), _safe(r.get("company_b")),
                    f"{(r.get('score') or 0):.1%}", RISK_NAMES.get(risk, risk),
                    _safe(r.get("summary"))]
            row = idx + 1
            _write_row(ws, row, vals)
            ws.cell(row=row, column=6).fill = RISK_COLORS.get(risk, PatternFill())
            if risk in ["critical", "high"]:
                ws.cell(row=row, column=6).font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 50

    # ==================== Sheet 4: 检测详情 (核心) ====================
    @staticmethod
    def _build_detail_sheet(wb, results):
        ws = wb.create_sheet("检测详情")
        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "各项预警详细检测数据"
        ws[f"A{row}"].font = TITLE_FONT
        ws.row_dimensions[row].height = 32
        row += 2

        sorted_results = sorted(results, key=lambda x: x.get("score", 0), reverse=True)

        for r_idx, r in enumerate(sorted_results):
            atype = r.get("analysis_type", "")
            details = r.get("details") or {}
            risk = r.get("risk_level") or "low"
            score = r.get("score") or 0

            # === Section header ===
            ws.merge_cells(f"A{row}:F{row}")
            title = f"【{TYPE_NAMES.get(atype, atype)}】{_safe(r.get('company_a'))} vs {_safe(r.get('company_b'))} — {score:.1%}（{RISK_NAMES.get(risk, risk)}）"
            ws[f"A{row}"] = title
            ws[f"A{row}"].font = SUBTITLE_FONT
            ws[f"A{row}"].fill = LIGHT_BLUE_FILL
            ws.row_dimensions[row].height = 24
            row += 1

            # Summary
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = _safe(r.get("summary"))
            ws[f"A{row}"].font = BODY_FONT
            row += 1

            # === Type-specific details ===
            if atype == "content_similarity":
                row = ExcelReportGenerator._write_content_sim_detail(ws, row, details)
            elif atype == "metadata_match":
                row = ExcelReportGenerator._write_alert_table(ws, row, details, "元数据匹配详情")
            elif atype == "format_match":
                row = ExcelReportGenerator._write_alert_table(ws, row, details, "格式一致项详情")
            elif atype == "timestamp_cluster":
                row = ExcelReportGenerator._write_timestamp_detail(ws, row, details)
            elif atype == "entity_cross":
                row = ExcelReportGenerator._write_entity_detail(ws, row, details)
            elif atype == "error_pattern":
                row = ExcelReportGenerator._write_error_pattern_detail(ws, row, details)
            elif atype == "price_analysis":
                row = ExcelReportGenerator._write_price_detail(ws, row, details)
            else:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = str(details)[:500]
                ws[f"A{row}"].font = DETAIL_FONT
                row += 1

            row += 1  # blank row between sections

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 22

    @staticmethod
    def _write_content_sim_detail(ws, row, details):
        # Sub-scores
        sims = [("SimHash", details.get("simhash_similarity")),
                ("TF-IDF余弦", details.get("cosine_similarity")),
                ("Jaccard", details.get("jaccard_similarity")),
                ("分词引擎", details.get("tokenizer"))]
        for label, val in sims:
            if val is not None:
                ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
                ws.cell(row=row, column=1).border = THIN_BORDER
                v = f"{val:.1%}" if isinstance(val, (int, float)) else str(val)
                ws.cell(row=row, column=2, value=v).font = BODY_FONT
                ws.cell(row=row, column=2).border = THIN_BORDER
                row += 1

        # Similar segments
        segments = details.get("similar_segments", [])
        if segments:
            row += 1
            headers = ["#", "相似度", "文档A段落", "文档B段落"]
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=i, value=h)
                c.font = HEADER_FONT; c.fill = SUB_HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
            row += 1
            for i, seg in enumerate(segments[:20], 1):
                sim = seg.get("similarity", 0)
                vals = [i, f"{sim:.0%}", _safe(seg.get("text_a_segment"))[:200], _safe(seg.get("text_b_segment"))[:200]]
                _write_row(ws, row, vals, aligns=[CENTER, CENTER, WRAP_TOP, WRAP_TOP])
                if sim >= 0.6:
                    for c in range(1, 5):
                        ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL
                ws.row_dimensions[row].height = max(30, min(60, len(vals[2]) // 3))
                row += 1
        return row

    @staticmethod
    def _write_alert_table(ws, row, details, title):
        alerts = details.get("alerts", [])
        if not alerts:
            return row
        headers = ["检测项", "文档A", "文档B", "说明"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = HEADER_FONT; c.fill = SUB_HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        row += 1
        for a in alerts:
            vals = [_safe(a.get("field") or a.get("type")),
                    str(_safe(a.get("value_a"), "-")),
                    str(_safe(a.get("value_b"), "-")),
                    _safe(a.get("description") or a.get("message"))]
            _write_row(ws, row, vals, aligns=[LEFT, LEFT, LEFT, LEFT])
            row += 1
        return row

    @staticmethod
    def _write_timestamp_detail(ws, row, details):
        info = [("聚集组数", details.get("cluster_count", 0)),
                ("时间窗口(分钟)", details.get("threshold_minutes", 5)),
                ("文档总数", details.get("total_documents", 0))]
        for label, val in info:
            ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=str(val)).font = BODY_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1
        clusters = details.get("clusters", [])
        for i, cl in enumerate(clusters):
            ws.merge_cells(f"A{row}:D{row}")
            docs_str = ", ".join([str(d.get("company") or d.get("file_name") or d) for d in cl.get("documents", cl.get("doc_ids", []))])
            ws[f"A{row}"] = f"聚集组{i+1}: {docs_str}"
            ws[f"A{row}"].font = DETAIL_FONT
            row += 1
        return row

    @staticmethod
    def _write_entity_detail(ws, row, details):
        hits = details.get("hits", [])
        if not hits:
            return row
        headers = ["实体类型", "泄露值", "说明"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = HEADER_FONT; c.fill = SUB_HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        row += 1
        for h in hits:
            vals = [_safe(h.get("type") or h.get("entity_type")),
                    _safe(h.get("value") or h.get("entity")),
                    _safe(h.get("description") or h.get("context"))]
            _write_row(ws, row, vals, aligns=[CENTER, LEFT, LEFT])
            row += 1
        return row

    @staticmethod
    def _write_error_pattern_detail(ws, row, details):
        items = []
        for e in details.get("common_errors", []):
            items.append(("共性错误", str(e) if isinstance(e, str) else _safe(e.get("error") or e.get("word"), str(e))))
        for s in details.get("common_standards", []):
            items.append(("过期标准", str(s) if isinstance(s, str) else _safe(s.get("standard") or s.get("code"), str(s))))
        for p in details.get("common_patterns", []):
            items.append(("共性模式", str(p) if isinstance(p, str) else _safe(p.get("pattern") or p.get("description"), str(p))))
        if details.get("type"):
            items.insert(0, ("错误类型", details["type"]))
        if not items:
            return row
        headers = ["类别", "内容"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = HEADER_FONT; c.fill = SUB_HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        row += 1
        for cat, content in items:
            _write_row(ws, row, [cat, content], aligns=[CENTER, LEFT])
            row += 1
        return row

    @staticmethod
    def _write_price_detail(ws, row, details):
        info = [("异常类型", _safe(details.get("alert_type")))]
        for key, label in [("arithmetic", "等差数列"), ("geometric", "等比数列"), ("fixed_coeff", "固定系数"), ("cluster", "价格聚集")]:
            sub = details.get(key, {})
            if sub and sub.get("detected"):
                info.append((label, f"已检出 — {str(sub)[:200]}"))
        for label, val in info:
            ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=str(val)).font = BODY_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = WRAP_TOP
            row += 1
        return row

    # ==================== Sheet 5: 风险预警汇总 ====================
    @staticmethod
    def _build_alerts_sheet(wb, results):
        ws = wb.create_sheet("风险预警汇总")
        type_counts = {}
        risk_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for r in results:
            t = r.get("analysis_type", "other")
            type_counts[t] = type_counts.get(t, 0) + 1
            rl = r.get("risk_level") or "low"
            if rl in risk_counts:
                risk_counts[rl] += 1

        ws["A1"] = "预警统计"; ws["A1"].font = SUBTITLE_FONT
        ws["A3"] = "风险等级"; ws["B3"] = "数量"
        ws["A3"].font = HEADER_FONT; ws["A3"].fill = HEADER_FILL
        ws["B3"].font = HEADER_FONT; ws["B3"].fill = HEADER_FILL
        for i, (level, count) in enumerate(risk_counts.items()):
            row = 4 + i
            ws.cell(row=row, column=1, value=RISK_NAMES[level]).font = BODY_FONT
            ws.cell(row=row, column=1).fill = RISK_COLORS[level]
            ws.cell(row=row, column=2, value=count).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = CENTER

        ws["A9"] = "检测类型"; ws["B9"] = "数量"
        ws["A9"].font = HEADER_FONT; ws["A9"].fill = HEADER_FILL
        ws["B9"].font = HEADER_FONT; ws["B9"].fill = HEADER_FILL
        for i, (t, count) in enumerate(type_counts.items()):
            row = 10 + i
            ws.cell(row=row, column=1, value=TYPE_NAMES.get(t, t)).font = BODY_FONT
            ws.cell(row=row, column=2, value=count).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = CENTER
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
